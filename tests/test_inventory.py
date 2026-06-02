import pytest
from openpyxl import load_workbook

from mia.core import inventory
from mia.core.common import Cancelled
from tests.conftest import CancelNow


@pytest.mark.parametrize("desc,expected", [
    ("Ax T1", "T1"),
    ("Ax T1 POST GAD", "T1+Gd"),
    ("T1 CON GADOLINIO", "T1+Gd"),          # Spanish post-contrast
    ("Ax FLAIR", "FLAIR"),
    ("DIFUSION", "DWI"),                     # Spanish diffusion
    ("3D TOF MRA", "MRA"),
    ("", "Unknown"),
    ("some random label", "Other"),
])
def test_detect_sequence(desc, expected):
    assert inventory.detect_sequence(desc) == expected


def test_scan_finds_studies_and_series(dataset_dir):
    root, expected = dataset_dir
    result = inventory.scan_directory(root)

    assert result.study_count == expected["studies"]
    assert result.errors == 0
    # Two distinct patient IDs across the studies
    ids = {s["patient_id"] for s in result.studies.values()}
    assert ids == expected["patient_ids"]

    # The study with three series should classify them correctly
    seqs = set()
    for study in result.studies.values():
        for series in study["series"].values():
            seqs.add(series["sequence_type"])
    assert "T1" in seqs and "T1+Gd" in seqs and "FLAIR" in seqs


def test_build_inventory_writes_three_sheets(dataset_dir, tmp_path):
    root, _ = dataset_dir
    out = tmp_path / "inv.xlsx"
    result = inventory.build_inventory(root, str(out))

    assert result.output_path == str(out)
    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == ["Studies", "Series Detail", "Consistency Check"]
    # One data row per study on the Studies sheet (plus header)
    assert wb["Studies"].max_row == 1 + result.study_count


def test_scan_reports_progress(dataset_dir):
    root, _ = dataset_dir
    events = []
    inventory.scan_directory(root, progress=events.append)
    scan_events = [e for e in events if e.phase == "scan"]
    assert scan_events
    assert scan_events[-1].done == scan_events[-1].total


def test_scan_cancellation(dataset_dir):
    root, _ = dataset_dir
    with pytest.raises(Cancelled):
        inventory.scan_directory(root, cancel=CancelNow())
