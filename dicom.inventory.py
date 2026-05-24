#!/usr/bin/env python3
"""
DICOM Inventory Script
======================

Walks a directory tree, finds all DICOM files (with or without .dcm extension),
groups them by study and series, and outputs an Excel inventory.

Designed for compiling years of hospital imaging CDs into one master view.

Usage:
    python3 dicom_inventory.py /path/to/raw_discs
    python3 dicom_inventory.py /path/to/raw_discs -o my_inventory.xlsx
    python3 dicom_inventory.py /path/to/raw_discs -v   (verbose progress)

Output:
    An .xlsx file with three sheets:
      1. Studies      - one row per study (date, modality, sequences present)
      2. Series Detail - one row per series (T1, T2, FLAIR, etc.)
      3. Consistency Check - lists all patient names/IDs/birthdates found,
         to flag if studies from different patients got mixed in

Install requirements:
    pip install pydicom openpyxl --break-system-packages
"""

import argparse
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import pydicom
    from pydicom.errors import InvalidDicomError
except ImportError:
    print("ERROR: pydicom not installed.")
    print("Run: pip install pydicom --break-system-packages")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("Run: pip install openpyxl --break-system-packages")
    sys.exit(1)


# Patterns to detect MR/CT sequence type from SeriesDescription.
# Includes English and Spanish keywords because hospital labels vary.
SEQUENCE_PATTERNS = {
    "T1+Gd":   [r"T1.*(?:GAD|GD|\+\s*C|POST|CONTRAS|CE\b)",
                r"(?:GAD|POST|CONTRAS).*T1",
                r"T1.*\+", r"POST.*GAD", r"T1.*C\+"],
    "T1":      [r"\bT1\b", r"MPRAGE", r"SPGR", r"TFE.*T1", r"BRAVO"],
    "T2":      [r"\bT2\b(?!\*)", r"TSE.*T2", r"FSE.*T2", r"PROPELLER.*T2"],
    "T2*":     [r"T2\*", r"\bGRE\b", r"\bSWI\b", r"SWAN", r"SUSCEPT"],
    "FLAIR":   [r"FLAIR", r"DARK.?FLUID"],
    "DWI":     [r"\bDWI\b", r"DIFF", r"DIFUSI", r"EPI.*DIFF"],
    "ADC":     [r"\bADC\b", r"APPARENT.*DIFF"],
    "MRA":     [r"\bMRA\b", r"\bTOF\b", r"TIME.?OF.?FLIGHT", r"ANGIO.*MR",
                r"3D.*TOF", r"ANGIO.*RM", r"ARTERIO"],
    "CTA":     [r"\bCTA\b", r"ANGIO.*CT", r"CT.*ANGIO", r"ANGIO.*TC"],
    "Perfusion":[r"PERF", r"\bPWI\b", r"\bDSC\b", r"\bDCE\b", r"PERFUSI"],
    "MRS":     [r"\bMRS\b", r"SPECTROS", r"ESPECTRO"],
    "Localizer":[r"LOCALIZ", r"SCOUT", r"SURVEY", r"TOPOGRAM", r"PLANO"],
    "Reformat":[r"REFORMAT", r"\bMPR\b", r"REFORMATED", r"RECON"],
}


def detect_sequence(series_desc):
    """Identify MR/CT sequence types from a series description string."""
    if not series_desc:
        return "Unknown"

    desc = series_desc.upper()
    matches = []
    for seq_name, patterns in SEQUENCE_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, desc):
                matches.append(seq_name)
                break

    # If we matched both T1 and T1+Gd, the post-contrast tag wins.
    if "T1+Gd" in matches and "T1" in matches:
        matches.remove("T1")

    return ", ".join(matches) if matches else "Other"


def safe_get(ds, attr, default=""):
    """Safely read a DICOM attribute; return default if missing/None."""
    val = getattr(ds, attr, default)
    if val is None or val == "":
        return default
    return str(val).strip()


def format_date(date_str):
    """Convert DICOM date YYYYMMDD -> YYYY-MM-DD."""
    if not date_str or len(date_str) < 8:
        return date_str
    return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"


def is_dicom_file(filepath):
    """Quick magic-byte check: real DICOM files have 'DICM' at offset 128."""
    try:
        with open(filepath, "rb") as f:
            f.seek(128)
            return f.read(4) == b"DICM"
    except (OSError, IOError):
        return False


def has_contrast(ds):
    """Detect contrast administration from DICOM tags."""
    agent = safe_get(ds, "ContrastBolusAgent")
    if agent:
        return agent
    route = safe_get(ds, "ContrastBolusRoute")
    if route:
        return f"Yes ({route})"
    return ""


# File extensions to skip without opening
SKIP_EXTENSIONS = {
    ".exe", ".dll", ".htm", ".html", ".jpg", ".jpeg", ".png", ".gif",
    ".pdf", ".txt", ".ini", ".inf", ".bat", ".css", ".js", ".xml",
    ".doc", ".docx", ".zip", ".log", ".db", ".plist",
}
SKIP_FILENAMES = {"DICOMDIR", "AUTORUN.INF", "README.TXT", "INDEX.HTM",
                  "INDEX.HTML", ".DS_STORE"}
SKIP_DIRS = {"viewer", "osirix", "autorun", "$recycle.bin", "system volume information"}


def scan_directory(root_dir, verbose=False):
    """Walk root_dir, parse DICOM headers, group by StudyInstanceUID."""
    studies = {}
    files_seen = 0
    dicoms_read = 0
    errors = 0

    print(f"Scanning {root_dir} ...")
    print("(this can take several minutes for large libraries)\n")

    for dirpath, dirnames, filenames in os.walk(root_dir):
        # Prune hidden and obviously-non-DICOM directories in-place
        dirnames[:] = [
            d for d in dirnames
            if not d.startswith(".") and d.lower() not in SKIP_DIRS
        ]

        for filename in filenames:
            files_seen += 1

            if filename.startswith(".") or filename.upper() in SKIP_FILENAMES:
                continue
            if Path(filename).suffix.lower() in SKIP_EXTENSIONS:
                continue

            filepath = os.path.join(dirpath, filename)

            if not is_dicom_file(filepath):
                continue

            try:
                ds = pydicom.dcmread(
                    filepath, stop_before_pixels=True, force=True
                )
            except (InvalidDicomError, Exception):
                errors += 1
                continue

            dicoms_read += 1

            study_uid = safe_get(ds, "StudyInstanceUID")
            series_uid = safe_get(ds, "SeriesInstanceUID")
            if not study_uid or not series_uid:
                continue

            if study_uid not in studies:
                studies[study_uid] = {
                    "study_date": format_date(safe_get(ds, "StudyDate")),
                    "study_time": safe_get(ds, "StudyTime")[:6],
                    "study_description": safe_get(ds, "StudyDescription"),
                    "modality": safe_get(ds, "Modality"),
                    "accession": safe_get(ds, "AccessionNumber"),
                    "patient_name": safe_get(ds, "PatientName"),
                    "patient_id": safe_get(ds, "PatientID"),
                    "patient_birth": format_date(
                        safe_get(ds, "PatientBirthDate")
                    ),
                    "patient_sex": safe_get(ds, "PatientSex"),
                    "patient_age": safe_get(ds, "PatientAge"),
                    "institution": safe_get(ds, "InstitutionName"),
                    "manufacturer": safe_get(ds, "Manufacturer"),
                    "model": safe_get(ds, "ManufacturerModelName"),
                    "field_strength": safe_get(ds, "MagneticFieldStrength"),
                    "kvp": safe_get(ds, "KVP"),
                    "series": {},
                    "source_path": dirpath,
                }

            if series_uid not in studies[study_uid]["series"]:
                series_desc = safe_get(ds, "SeriesDescription")
                studies[study_uid]["series"][series_uid] = {
                    "series_number": safe_get(ds, "SeriesNumber"),
                    "series_description": series_desc,
                    "modality": safe_get(ds, "Modality"),
                    "sequence_type": detect_sequence(series_desc),
                    "body_part": safe_get(ds, "BodyPartExamined"),
                    "contrast": has_contrast(ds),
                    "slice_thickness": safe_get(ds, "SliceThickness"),
                    "pixel_spacing": safe_get(ds, "PixelSpacing"),
                    "image_count": 0,
                    "source_path": dirpath,
                }

            studies[study_uid]["series"][series_uid]["image_count"] += 1

            if verbose and dicoms_read % 500 == 0:
                print(f"  ... {dicoms_read} DICOM files processed")

    print(f"\nScan complete:")
    print(f"  Files examined : {files_seen}")
    print(f"  DICOM files    : {dicoms_read}")
    print(f"  Read errors    : {errors}")
    print(f"  Studies found  : {len(studies)}")
    return studies


def write_inventory_xlsx(studies, output_path):
    """Build the three-sheet Excel inventory."""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C5282",
                              end_color="2C5282", fill_type="solid")
    alt_fill = PatternFill(start_color="F7FAFC",
                           end_color="F7FAFC", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # ----- Sheet 1: Studies -----
    ws1 = wb.active
    ws1.title = "Studies"

    headers = [
        "#", "Study Date", "Modality", "Description", "Body Part",
        "# Series", "# Images", "T1", "T2", "FLAIR", "DWI", "T1+Gd",
        "MRA/CTA", "Institution", "Scanner", "Field/kVp",
        "Accession", "Source Folder",
    ]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    sorted_studies = sorted(
        studies.items(), key=lambda x: x[1].get("study_date", "")
    )

    for idx, (_, study) in enumerate(sorted_studies, 1):
        seq_set = set()
        body_parts = set()
        total_imgs = 0
        for s in study["series"].values():
            for tag in s["sequence_type"].split(", "):
                seq_set.add(tag)
            if s["body_part"]:
                body_parts.add(s["body_part"])
            total_imgs += s["image_count"]

        # Build the field/kVp string per modality
        if study["modality"] == "MR" and study["field_strength"]:
            field_or_kvp = f"{study['field_strength']}T"
        elif study["modality"] == "CT" and study["kvp"]:
            field_or_kvp = f"{study['kvp']} kVp"
        else:
            field_or_kvp = ""

        scanner = f"{study['manufacturer']} {study['model']}".strip()
        source = os.path.basename(study["source_path"]) or study["source_path"]

        row = [
            idx,
            study["study_date"],
            study["modality"],
            study["study_description"],
            ", ".join(sorted(body_parts)),
            len(study["series"]),
            total_imgs,
            "✓" if "T1" in seq_set else "",
            "✓" if "T2" in seq_set else "",
            "✓" if "FLAIR" in seq_set else "",
            "✓" if ("DWI" in seq_set or "ADC" in seq_set) else "",
            "✓" if "T1+Gd" in seq_set else "",
            "✓" if ("MRA" in seq_set or "CTA" in seq_set) else "",
            study["institution"],
            scanner,
            field_or_kvp,
            study["accession"],
            source,
        ]
        for col, val in enumerate(row, 1):
            c = ws1.cell(row=idx + 1, column=col, value=val)
            if idx % 2 == 0:
                c.fill = alt_fill
            if 8 <= col <= 13:
                c.alignment = center

    col_widths_1 = [4, 12, 9, 35, 18, 9, 10, 6, 6, 7, 6, 7, 9,
                    25, 24, 12, 14, 22]
    for i, w in enumerate(col_widths_1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # ----- Sheet 2: Series Detail -----
    ws2 = wb.create_sheet("Series Detail")
    headers2 = [
        "Study #", "Study Date", "Modality", "Series #", "Description",
        "Sequence Type", "Body Part", "Contrast", "# Images",
        "Slice Thk", "Pixel Spacing", "Source Folder",
    ]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    row = 2
    for idx, (_, study) in enumerate(sorted_studies, 1):
        def _sort_key(item):
            sn = item[1]["series_number"]
            return int(sn) if sn.isdigit() else 999

        for _, series in sorted(study["series"].items(), key=_sort_key):
            source = os.path.basename(series["source_path"]) or series["source_path"]
            row_data = [
                idx,
                study["study_date"],
                series["modality"],
                series["series_number"],
                series["series_description"],
                series["sequence_type"],
                series["body_part"],
                series["contrast"],
                series["image_count"],
                series["slice_thickness"],
                series["pixel_spacing"],
                source,
            ]
            for col, val in enumerate(row_data, 1):
                c = ws2.cell(row=row, column=col, value=val)
                if idx % 2 == 0:
                    c.fill = alt_fill
            row += 1

    col_widths_2 = [8, 12, 9, 8, 38, 22, 14, 22, 10, 10, 18, 22]
    for i, w in enumerate(col_widths_2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ----- Sheet 3: Consistency Check -----
    ws3 = wb.create_sheet("Consistency Check")
    ws3.cell(row=1, column=1,
             value="Patient Consistency Check").font = Font(bold=True, size=14)
    ws3.cell(row=2, column=1,
             value=("If multiple values appear in any group, you may have "
                    "studies from different patients mixed together, OR the "
                    "same patient registered differently across hospitals."))

    def _dump_group(title, counts, start_row):
        ws3.cell(row=start_row, column=1, value=title).font = Font(bold=True)
        r = start_row + 1
        for value, count in sorted(counts.items(), key=lambda x: -x[1]):
            ws3.cell(row=r, column=1, value=value)
            ws3.cell(row=r, column=2, value=f"{count} studies")
            r += 1
        return r + 1

    pids = defaultdict(int)
    pnames = defaultdict(int)
    pbirths = defaultdict(int)
    for study in studies.values():
        if study["patient_id"]:
            pids[study["patient_id"]] += 1
        if study["patient_name"]:
            pnames[study["patient_name"]] += 1
        if study["patient_birth"]:
            pbirths[study["patient_birth"]] += 1

    r = 4
    r = _dump_group("Patient IDs found:", pids, r)
    r = _dump_group("Patient Names found:", pnames, r)
    r = _dump_group("Birth dates found:", pbirths, r)

    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["B"].width = 22

    wb.save(output_path)
    print(f"\nInventory saved to: {output_path}")


def print_summary(studies):
    """Console summary so the user gets immediate feedback."""
    if not studies:
        print("\nNo DICOM studies found.")
        return

    sorted_studies = sorted(
        studies.items(), key=lambda x: x[1].get("study_date", "")
    )

    print("\n" + "=" * 72)
    print("INVENTORY SUMMARY")
    print("=" * 72)

    for idx, (_, study) in enumerate(sorted_studies, 1):
        seq_set = set()
        total_imgs = 0
        for s in study["series"].values():
            seq_set.update(s["sequence_type"].split(", "))
            total_imgs += s["image_count"]
        seq_set -= {"Unknown", "Other", "Localizer", "Reformat", ""}

        desc = (study["study_description"] or "")[:42]
        print(f"\n[{idx:2d}] {study['study_date']} | {study['modality']:3s} | {desc}")
        if study["institution"]:
            print(f"     Institution : {study['institution'][:55]}")
        print(f"     Series      : {len(study['series'])}   Images: {total_imgs}")
        if seq_set:
            print(f"     Sequences   : {', '.join(sorted(seq_set))}")

    print("\n" + "=" * 72)
    print(f"Total: {len(studies)} studies")
    print("=" * 72)


def main():
    parser = argparse.ArgumentParser(
        description="Build a master inventory of all DICOM studies under a directory.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=("Examples:\n"
                "  python3 dicom_inventory.py ~/Imaging_Master/raw_discs\n"
                "  python3 dicom_inventory.py ~/Imaging_Master/raw_discs -o my_inventory.xlsx\n"
                "  python3 dicom_inventory.py ~/Imaging_Master/raw_discs -v\n"),
    )
    parser.add_argument("directory", help="Root directory to scan")
    parser.add_argument("-o", "--output", default="dicom_inventory.xlsx",
                        help="Output xlsx filename (default: dicom_inventory.xlsx)")
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="Show progress every 500 files")

    args = parser.parse_args()

    root = os.path.expanduser(args.directory)
    if not os.path.isdir(root):
        print(f"ERROR: {root} is not a directory")
        sys.exit(1)

    studies = scan_directory(root, verbose=args.verbose)
    print_summary(studies)
    if studies:
        write_inventory_xlsx(studies, args.output)


if __name__ == "__main__":
    main()